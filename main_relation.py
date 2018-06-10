import openpyxl
import numpy as np
import math
import matplotlib.pyplot as plt
from numpy.linalg import inv

alarmdata=openpyxl.load_workbook('C:\\Users\\aviza\\Desktop\\CSE 674\\Project 1\\alarm10K.xlsx')
aladata=alarmdata.get_sheet_by_name('alarm10K')
individualp=[]

def find_pdf(a,b):
    facts=1
    probs=1
    for i in range(len(a)):
        facts*=math.factorial(int(b[i]))
        probs*=math.pow(a[i],int(b[i]))
          
    pd=math.factorial(100)*probs/(facts)
    return pd/10000

def find_info(x,y):
    fx=[]
    fy=[]
    for i in range(2,10001,1):
        fx.append(aladata.cell(row=i,column=x).value)
        fy.append(aladata.cell(row=i,column=y).value)
        
    x1=list(set(fx))
    y1=list(set(fy))
    info=0   
    for q in x1:
        
        for r in y1:
            s=0
            
            for j in range(2,10001,1):
                if aladata.cell(row=j,column=x).value == q and aladata.cell(row=j,column=y).value == r:
                    s+=1

                
            s=s/10000
            p1=find_count(fx,q)
            p2=find_count(fy,r)
            #try:
            info+=s*math.log(  (s/(p1*p2))  +1 )

            #except:
                #pass

    return info
                    
def find_count(li,v):
    sq=0.0
    for i in range(len(li)):
        if li[i]==v:
            sq+=1

    return sq/len(li)
    
    

for k in range(1,38,1):
    l=[]
    for i in range(2,10001,1):
        l.append(aladata.cell(row=i,column=k).value)
    probab=list(set(l))
   

    s=[]
    outcomes=[]
    for p in probab:
        sum1=0
        for l1 in l:
            if l1==p:
                sum1+=1
        outcomes.append(sum1/100.0)
        s.append(sum1/10000.0)

    
    individualp.append(find_pdf(s,outcomes))

mat= np.zeros((37,37))
rel_mat = np.zeros((37,37))
count =0;


for mi in range(1,38,1):
    for mj in range(1,38,1):
        
        if mi==mj:
            mat[mi-1][mj-1]=1
        if mi<mj:
            mat[mi-1][mj-1]=find_info(mi,mj)
            if ( mat[mi-1][mj-1] > 0.7):
                rel_mat[mi-1][mj-1] = 1
                count = count + 1        
           
print(mat)
print (rel_mat)
print (count)